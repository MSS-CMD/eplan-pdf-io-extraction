# -*- coding: utf-8 -*-
"""
WV26007-徐工模组&PACK装配线L2_PLC3-电芯包膜-V1
PLC I/O表提取脚本 + 设备清单（含安全输出）
"""
import pymupdf, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter

PDF_PATH = r'path\to\your\drawing.pdf'
OUTPUT_PATH = r'PLC_IO_Table.xlsx'

doc = pymupdf.open(PDF_PATH)

def get_text_items(page):
    items = []
    for b in page.get_text("dict")["blocks"]:
        if b["type"] == 0:
            for line in b["lines"]:
                text = "".join([s["text"] for s in line["spans"]]).strip()
                if text:
                    x0, y0, x1, y1 = line["bbox"]
                    items.append((text, x0, y0, x1, y1))
    return items

def in_column(x_center, col_left, col_right):
    return col_left <= x_center <= col_right

def unique_desc(lines):
    seen = set()
    out = []
    for d in lines:
        if d not in seen:
            seen.add(d)
            out.append(d)
    return " ".join(out).strip()

# ==================== DI/安全输入解析 ====================

def parse_di_safety(page, page_num, io_type="DI"):
    items = get_text_items(page)
    records = []
    module_name = ""
    for text, x0, y0, x1, y1 in items:
        if "数字量输入模块" in text or "安全输入模块" in text:
            module_name = text.strip()
    anchors = []
    for text, x0, y0, x1, y1 in items:
        m = re.match(r"^(DI|FDI)\s+Bit\s+(\d+)$", text)
        if m:
            cx = (x0 + x1) / 2
            anchors.append((cx, int(m.group(2))))
    if not anchors:
        return records
    anchors.sort(key=lambda a: a[0])
    cols = []
    for idx, (cx, bit) in enumerate(anchors):
        left = (anchors[idx-1][0] + cx)/2 if idx > 0 else cx - 60
        right = (anchors[idx+1][0] + cx)/2 if idx < len(anchors)-1 else cx + 60
        cols.append((left, right, cx, bit))
    for left, right, cx, bit in cols:
        address = ""
        device_tag = ""
        desc_lines = []
        for text, x0, y0, x1, y1 in items:
            xc = (x0 + x1) / 2
            if not in_column(xc, left, right):
                continue
            if re.match(r"^I\d{4}\.\d$", text) and 580 < y0 < 605:
                address = text
            if (text.startswith("+") or text.startswith("-")) and 590 < y0 < 615:
                device_tag = text
            if re.search(r'[\u4e00-\u9fff]', text) and 665 < y0 < 715 and "Bit" not in text:
                desc_lines.append(text)
            if text == "备用" and 675 < y0 < 710:
                desc_lines.append("备用")
        if address:
            records.append({
                "address": address,
                "device_tag": device_tag,
                "description": unique_desc(desc_lines),
                "module": module_name,
                "io_type": io_type,
                "page": page_num,
                "bit": bit
            })
    return records

# ==================== DQ解析 ====================

def parse_dq(page, page_num):
    items = get_text_items(page)
    records = []
    module_name = ""
    for text, x0, y0, x1, y1 in items:
        if "数字量输出模块" in text:
            module_name = text.strip()
    anchors = []
    for text, x0, y0, x1, y1 in items:
        m = re.match(r"^DO\s+Bit\s+(\d+)$", text)
        if m:
            cx = (x0 + x1) / 2
            anchors.append((cx, int(m.group(1))))
    if not anchors:
        return records
    anchors.sort(key=lambda a: a[0])
    cols = []
    for idx, (cx, bit) in enumerate(anchors):
        left = (anchors[idx-1][0] + cx)/2 if idx > 0 else cx - 60
        right = (anchors[idx+1][0] + cx)/2 if idx < len(anchors)-1 else cx + 60
        cols.append((left, right, cx, bit))
    for left, right, cx, bit in cols:
        address = ""
        device_tag = ""
        desc_lines = []
        for text, x0, y0, x1, y1 in items:
            xc = (x0 + x1) / 2
            if not in_column(xc, left, right):
                continue
            if re.match(r"^Q\d{4}\.\d$", text) and 148 < y0 < 172:
                address = text
            if (text.startswith("+") or text.startswith("-")) and 140 < y0 < 160:
                device_tag = text
            if re.search(r'[\u4e00-\u9fff]', text) and 665 < y0 < 715 and "Bit" not in text:
                desc_lines.append(text)
            if text == "备用" and 675 < y0 < 710:
                desc_lines.append("备用")
        if address:
            records.append({
                "address": address,
                "device_tag": device_tag,
                "description": unique_desc(desc_lines),
                "module": module_name,
                "io_type": "DQ",
                "page": page_num,
                "bit": bit
            })
    return records

# ==================== 安全输出解析 ====================

def parse_safety_output(page, page_num):
    """
    安全输出模块布局（类似DQ但标签为DQ-P0/DQ-P1等）
    """
    items = get_text_items(page)
    records = []
    module_name = ""
    for text, x0, y0, x1, y1 in items:
        if "安全输出模块" in text:
            module_name = text.strip()
    # 找DQ-P标签作为列锚点
    anchors = []
    for text, x0, y0, x1, y1 in items:
        m = re.match(r"^DQ-P\d+\.?$", text)
        if m and 170 < y0 < 190:
            cx = (x0 + x1) / 2
            anchors.append((cx, text.strip()))
    if not anchors:
        return records
    anchors.sort(key=lambda a: a[0])
    cols = []
    for idx, (cx, label) in enumerate(anchors):
        left = (anchors[idx-1][0] + cx)/2 if idx > 0 else cx - 60
        right = (anchors[idx+1][0] + cx)/2 if idx < len(anchors)-1 else cx + 60
        cols.append((left, right, cx, label))
    for left, right, cx, label in cols:
        address = ""
        desc_lines = []
        for text, x0, y0, x1, y1 in items:
            xc = (x0 + x1) / 2
            if not in_column(xc, left, right):
                continue
            if re.match(r"^Q1110\.\d$", text) and 148 < y0 < 172:
                address = text
            if re.search(r'[\u4e00-\u9fff]', text) and 665 < y0 < 715 and "Bit" not in text and "DQ-P" not in text:
                desc_lines.append(text)
            if text == "备用" and 675 < y0 < 710:
                desc_lines.append("备用")
        if address:
            records.append({
                "address": address,
                "device_tag": "",
                "description": unique_desc(desc_lines),
                "module": module_name,
                "io_type": "安全输出(FDQ)",
                "page": page_num,
                "bit": label
            })
    return records

# ==================== 主循环 ====================

all_records = []

print("正在解析DI页...")
for pn in range(86, 98):
    records = parse_di_safety(doc[pn-1], pn, "DI")
    all_records.extend(records)
    print("  第{}页: {} 个DI点".format(pn, len(records)))

print("正在解析DQ页...")
for pn in range(98, 104):
    records = parse_dq(doc[pn-1], pn)
    all_records.extend(records)
    print("  第{}页: {} 个DQ点".format(pn, len(records)))

print("正在解析安全输入页...")
for pn in [75, 77, 79, 81, 83]:
    records = parse_di_safety(doc[pn-1], pn, "安全输入(FDI)")
    all_records.extend(records)
    print("  第{}页: {} 个安全输入点".format(pn, len(records)))

print("正在解析安全输出页...")
pn = 85
records = parse_safety_output(doc[pn-1], pn)
all_records.extend(records)
print("  第{}页: {} 个安全输出点".format(pn, len(records)))

doc.close()

# ==================== 排序 ====================

def addr_sort_key(r):
    a = r["address"]
    m = re.match(r"I(\d{4})\.(\d)", a)
    if m: return (0, int(m.group(1)), int(m.group(2)))
    m = re.match(r"Q(\d{4})\.(\d)", a)
    if m: return (1, int(m.group(1)), int(m.group(2)))
    return (9, 0, 0)

all_records.sort(key=addr_sort_key)

print("\n总I/O点数:", len(all_records))

# ==================== 设备清单 ====================

print("\n正在梳理设备清单...")
equipment_list = []
device_groups = {}

for rec in all_records:
    desc = rec["description"]
    tag = rec["device_tag"]
    addr = rec["address"]
    io_type = rec["io_type"]
    if "备用" in desc and len(desc.strip()) <= 4:
        continue
    primary_desc = desc.split(" ")[0] if desc else ""
    equipment_list.append({
        "device": primary_desc, "tag": tag, "address": addr,
        "io_type": io_type, "description": desc
    })
    if primary_desc and primary_desc not in ["备用"]:
        if primary_desc not in device_groups:
            device_groups[primary_desc] = {"DI": 0, "DQ": 0, "FDI": 0, "FDQ": 0, "addresses": [], "descriptions": []}
        type_key = io_type
        if io_type == "DI": device_groups[primary_desc]["DI"] += 1
        elif io_type == "DQ": device_groups[primary_desc]["DQ"] += 1
        elif io_type == "安全输入(FDI)": device_groups[primary_desc]["FDI"] += 1
        elif io_type == "安全输出(FDQ)": device_groups[primary_desc]["FDQ"] += 1
        device_groups[primary_desc]["addresses"].append(addr)
        if desc not in device_groups[primary_desc]["descriptions"]:
            device_groups[primary_desc]["descriptions"].append(desc)

print("\n共发现 {} 种设备类型".format(len(device_groups)))
print("\n========== 设备清单 ==========")
print("{:4s} | {:25s} | {:4s} | {:4s} | {:4s} | {:4s} | {:30s}".format("序号", "设备名称", "DI", "DQ", "FDI", "FDQ", "IO地址"))
print("-" * 95)

dev_sorted = sorted(device_groups.items(), key=lambda x: -(x[1]["DI"]+x[1]["DQ"]+x[1]["FDI"]+x[1]["FDQ"]))
for i, (dev, info) in enumerate(dev_sorted, 1):
    addrs = ", ".join(info["addresses"][:5])
    if len(info["addresses"]) > 5:
        addrs += "..."
    print("{:4d} | {:25s} | {:4d} | {:4d} | {:4d} | {:4d} | {:30s}".format(
        i, dev[:25], info["DI"], info["DQ"], info["FDI"], info["FDQ"], addrs
    ))

# ==================== 生成Excel ====================

wb = Workbook()
ws = wb.active
ws.title = "PLC IO 总表"

header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
data_font = Font(name='微软雅黑', size=10)
data_align = Alignment(vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)
type_fills = {
    "DI": PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid'),
    "DQ": PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
    "安全输入(FDI)": PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
    "安全输出(FDQ)": PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),
}

headers = ["序号", "I/O类型", "地址", "Bit", "描述", "设备标签", "模块/机架", "页码"]
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align

for idx, rec in enumerate(all_records, 1):
    row = idx + 1
    fill = type_fills.get(rec["io_type"], PatternFill())
    ws.cell(row=row, column=1, value=idx).font = data_font
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row, column=2, value=rec["io_type"]).font = data_font
    ws.cell(row=row, column=2).fill = fill
    ws.cell(row=row, column=2).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row, column=3, value=rec["address"]).font = Font(name='Consolas', size=10, bold=True)
    ws.cell(row=row, column=3).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row, column=4, value=rec["bit"]).font = data_font
    ws.cell(row=row, column=4).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row, column=5, value=rec["description"]).font = data_font
    ws.cell(row=row, column=5).alignment = data_align
    ws.cell(row=row, column=6, value=rec["device_tag"]).font = data_font
    ws.cell(row=row, column=6).alignment = data_align
    ws.cell(row=row, column=7, value=rec["module"]).font = data_font
    ws.cell(row=row, column=7).alignment = data_align
    ws.cell(row=row, column=8, value=rec["page"]).font = data_font
    ws.cell(row=row, column=8).alignment = Alignment(horizontal='center', vertical='center')
    for col in range(1, 9):
        ws.cell(row=row, column=col).border = thin_border

col_widths = [6, 16, 12, 8, 35, 40, 55, 8]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = "A1:H{}".format(len(all_records) + 1)

# 汇总统计表
ws2 = wb.create_sheet(title="汇总统计")
type_counter = Counter(r["io_type"] for r in all_records)
ws2.cell(row=1, column=1, value="I/O类型").font = Font(name='微软雅黑', bold=True, size=11)
ws2.cell(row=1, column=2, value="点数").font = Font(name='微软雅黑', bold=True, size=11)
ws2.cell(row=1, column=3, value="地址范围").font = Font(name='微软雅黑', bold=True, size=11)
cr = 2
total = 0
for io_type in ["DI", "DQ", "安全输入(FDI)", "安全输出(FDQ)"]:
    count = type_counter.get(io_type, 0)
    total += count
    recs = [r for r in all_records if r["io_type"] == io_type]
    addr_range = ""
    if recs:
        addr_range = "{} ~ {}".format(recs[0]["address"], recs[-1]["address"])
    ws2.cell(row=cr, column=1, value=io_type).font = data_font
    ws2.cell(row=cr, column=2, value=count).font = data_font
    ws2.cell(row=cr, column=3, value=addr_range).font = data_font
    cr += 1
ws2.cell(row=cr, column=1, value="合计").font = Font(name='微软雅黑', bold=True, size=11)
ws2.cell(row=cr, column=2, value=total).font = Font(name='微软雅黑', bold=True, size=11)
ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 10
ws2.column_dimensions['C'].width = 30

# 设备清单表
ws3 = wb.create_sheet(title="设备清单")
eq_headers = ["序号", "设备名称", "DI", "DQ", "安全输入(FDI)", "安全输出(FDQ)", "合计", "IO地址", "详细描述"]
for col_idx, h in enumerate(eq_headers, 1):
    cell = ws3.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align

for i, (dev, info) in enumerate(dev_sorted, 1):
    row = i + 1
    total_pts = info["DI"] + info["DQ"] + info["FDI"] + info["FDQ"]
    addrs = ", ".join(info["addresses"])
    descs = "; ".join(info["descriptions"])
    ws3.cell(row=row, column=1, value=i).font = data_font
    ws3.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws3.cell(row=row, column=2, value=dev).font = Font(name='微软雅黑', size=10, bold=True)
    ws3.cell(row=row, column=3, value=info["DI"]).font = data_font
    ws3.cell(row=row, column=3).alignment = Alignment(horizontal='center', vertical='center')
    ws3.cell(row=row, column=4, value=info["DQ"]).font = data_font
    ws3.cell(row=row, column=4).alignment = Alignment(horizontal='center', vertical='center')
    ws3.cell(row=row, column=5, value=info["FDI"]).font = data_font
    ws3.cell(row=row, column=5).alignment = Alignment(horizontal='center', vertical='center')
    ws3.cell(row=row, column=6, value=info["FDQ"]).font = data_font
    ws3.cell(row=row, column=6).alignment = Alignment(horizontal='center', vertical='center')
    ws3.cell(row=row, column=7, value=total_pts).font = Font(name='微软雅黑', size=10, bold=True)
    ws3.cell(row=row, column=7).alignment = Alignment(horizontal='center', vertical='center')
    ws3.cell(row=row, column=8, value=addrs).font = data_font
    ws3.cell(row=row, column=8).alignment = data_align
    ws3.cell(row=row, column=9, value=descs).font = data_font
    ws3.cell(row=row, column=9).alignment = data_align
    for col in range(1, 10):
        ws3.cell(row=row, column=col).border = thin_border

ws3.column_dimensions['A'].width = 6
ws3.column_dimensions['B'].width = 25
ws3.column_dimensions['C'].width = 6
ws3.column_dimensions['D'].width = 6
ws3.column_dimensions['E'].width = 14
ws3.column_dimensions['F'].width = 14
ws3.column_dimensions['G'].width = 6
ws3.column_dimensions['H'].width = 50
ws3.column_dimensions['I'].width = 60
ws3.freeze_panes = "A2"
ws3.auto_filter.ref = "A1:I{}".format(len(dev_sorted) + 1)

wb.save(OUTPUT_PATH)
print("\nExcel已保存至:", OUTPUT_PATH)
print("包含3个工作表: PLC IO总表, 汇总统计, 设备清单")

# 预览
print("\n========== 前10行预览 ==========")
print("序号  | 类型        | 地址      | 描述")
print("-" * 55)
for i, rec in enumerate(all_records[:10]):
    desc = rec["description"][:25] + ".." if len(rec["description"]) > 25 else rec["description"]
    print("{:4d} | {:12s} | {:9s} | {:25s}".format(
        i+1, rec["io_type"], rec["address"], desc
    ))
print("\n... (共 {} 条记录)".format(len(all_records)))
