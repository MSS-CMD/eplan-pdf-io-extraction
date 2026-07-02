# -*- coding: utf-8 -*-
"""
framework/utils/excel.py
Excel导出格式化
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 颜色方案
TYPE_FILLS = {
    "DI": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
    "DQ": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "安全输入(FDI)": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "安全输出(FDQ)": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "DIO (IOLINK-I)": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "DIO (IOLINK-Q)": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "阀岛(FDQ)": PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid"),
    "AI": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
}

HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
DATA_FONT = Font(name="微软雅黑", size=10)
ADDR_FONT = Font(name="Consolas", size=10, bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def generate_excel(records, output_path, title="PLC IO 总表"):
    """生成带格式的Excel"""
    wb = Workbook()
    # ===== Sheet1: IO总表 =====
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name max 31 chars

    # 自动检测列
    has_bit = any(r.get("bit") is not None and r["bit"] != "" for r in records[:10])
    has_tag = any(r.get("device_tag") for r in records[:10])

    if has_bit and has_tag:
        headers = ["序号", "I/O类型", "地址", "Bit", "描述", "设备标签", "模块/机架", "页码"]
        col_widths = [6, 16, 12, 8, 40, 40, 55, 8]
    elif has_bit:
        headers = ["序号", "I/O类型", "地址", "Bit", "描述", "模块/机架", "页码"]
        col_widths = [6, 16, 12, 8, 50, 55, 8]
    else:
        headers = ["序号", "I/O类型", "地址", "描述", "模块/机架", "页码"]
        col_widths = [6, 16, 12, 50, 55, 8]

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, rec in enumerate(records, 1):
        row = idx + 1
        fill = TYPE_FILLS.get(rec.get("io_type", ""), PatternFill())
        col = 1
        ws.cell(row=row, column=col, value=idx).font = DATA_FONT
        ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")
        col += 1
        ws.cell(row=row, column=col, value=rec.get("io_type", "")).font = DATA_FONT
        ws.cell(row=row, column=col).fill = fill
        ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")
        col += 1
        ws.cell(row=row, column=col, value=rec.get("address", "")).font = ADDR_FONT
        ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")
        col += 1
        if has_bit:
            ws.cell(row=row, column=col, value=rec.get("bit", "")).font = DATA_FONT
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")
            col += 1
        ws.cell(row=row, column=col, value=rec.get("description", "")).font = DATA_FONT
        ws.cell(row=row, column=col).alignment = Alignment(vertical="center", wrap_text=True)
        col += 1
        if has_tag:
            ws.cell(row=row, column=col, value=rec.get("device_tag", "")).font = DATA_FONT
            ws.cell(row=row, column=col).alignment = Alignment(vertical="center", wrap_text=True)
            col += 1
        ws.cell(row=row, column=col, value=str(rec.get("module", "")).strip()).font = DATA_FONT
        ws.cell(row=row, column=col).alignment = Alignment(vertical="center", wrap_text=True)
        col += 1
        ws.cell(row=row, column=col, value=rec.get("page", "")).font = DATA_FONT
        ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")

        for ci in range(1, len(headers) + 1):
            ws.cell(row=row, column=ci).border = THIN_BORDER

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:{}".format(get_column_letter(len(headers))) + str(len(records) + 1)

    # ===== Sheet2: 汇总 =====
    ws2 = wb.create_sheet(title="汇总统计")
    from collections import Counter
    type_counter = Counter(r.get("io_type", "") for r in records)
    ws2.cell(row=1, column=1, value="I/O类型").font = Font(name="微软雅黑", bold=True, size=11)
    ws2.cell(row=1, column=2, value="点数").font = Font(name="微软雅黑", bold=True, size=11)
    ws2.cell(row=1, column=3, value="地址范围").font = Font(name="微软雅黑", bold=True, size=11)

    cr = 2
    total = 0
    for io_type, count in type_counter.most_common():
        if count == 0:
            continue
        total += count
        recs = [r for r in records if r.get("io_type") == io_type]
        addr_range = ""
        if recs:
            addr_range = "{} ~ {}".format(recs[0]["address"], recs[-1]["address"])
        ws2.cell(row=cr, column=1, value=io_type).font = DATA_FONT
        ws2.cell(row=cr, column=2, value=count).font = DATA_FONT
        ws2.cell(row=cr, column=3, value=addr_range).font = DATA_FONT
        cr += 1
    ws2.cell(row=cr, column=1, value="合计").font = Font(name="微软雅黑", bold=True, size=11)
    ws2.cell(row=cr, column=2, value=total).font = Font(name="微软雅黑", bold=True, size=11)
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 35

    wb.save(output_path)
    return output_path
